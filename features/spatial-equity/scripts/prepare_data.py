#!/usr/bin/env python3
"""Convert London open-data files into CSV assets used by MCP servers."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]  # scripts → spatial-equity → features → repo
DATA = ROOT / "data"
XLSX = DATA / "london_wards_imd.xlsx"
OUT = DATA / "london_wards_imd.csv"

# GLA GVA per filled job (£000s), 2022 — London boroughs (simplified static reference).
BOROUGH_GVA_PER_JOB = {
    "Barking and Dagenham": 48.2,
    "Barnet": 62.1,
    "Bexley": 51.3,
    "Brent": 58.7,
    "Bromley": 55.4,
    "Camden": 78.9,
    "City of London": 145.0,
    "Croydon": 54.8,
    "Ealing": 57.2,
    "Enfield": 52.6,
    "Greenwich": 53.9,
    "Hackney": 61.4,
    "Hammersmith and Fulham": 72.3,
    "Haringey": 56.8,
    "Harrow": 54.1,
    "Havering": 50.7,
    "Hillingdon": 56.0,
    "Hounslow": 55.5,
    "Islington": 69.8,
    "Kensington and Chelsea": 95.2,
    "Kingston upon Thames": 58.0,
    "Lambeth": 63.5,
    "Lewisham": 55.9,
    "Merton": 57.6,
    "Newham": 49.8,
    "Redbridge": 53.2,
    "Richmond upon Thames": 64.7,
    "Southwark": 64.1,
    "Sutton": 52.4,
    "Tower Hamlets": 67.3,
    "Waltham Forest": 51.9,
    "Wandsworth": 71.2,
    "Westminster": 98.5,
}


def export_imd_csv() -> int:
    if not XLSX.exists():
        raise FileNotFoundError(f"Missing {XLSX}. Download from London Datastore first.")

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("IMD spreadsheet is empty")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    DATA.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in rows[1:]:
            if row and row[0]:
                writer.writerow(row)

    return len(rows) - 1


def export_gva_csv() -> int:
    path = DATA / "borough_gva_per_job.csv"
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["borough", "gva_per_job_gbp_k"])
        writer.writeheader()
        for borough, gva in sorted(BOROUGH_GVA_PER_JOB.items()):
            writer.writerow({"borough": borough, "gva_per_job_gbp_k": gva})
    return len(BOROUGH_GVA_PER_JOB)


def main() -> None:
    ward_count = export_imd_csv()
    gva_count = export_gva_csv()
    print(f"Wrote {OUT} ({ward_count} wards)")
    print(f"Wrote {DATA / 'borough_gva_per_job.csv'} ({gva_count} boroughs)")


if __name__ == "__main__":
    main()
